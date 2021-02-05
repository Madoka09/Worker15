'''
credits views
'''
from __future__ import unicode_literals

# Django
from django.conf import settings
from django.contrib.auth.decorators import user_passes_test
from django.core.files.storage import FileSystemStorage
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Q, Sum, Min, Max, Value
from django.db.models.functions import Concat
from django.http import JsonResponse
from django.shortcuts import render, redirect, get_list_or_404, HttpResponse
from django.template.loader import render_to_string
from django.views import View

#Python
from array import array
from datetime import datetime, timezone
from os import listdir
from os.path import isfile, join
from openpyxl import Workbook
from PIL import Image
import img2pdf, io, json, os

# Forms
#from credits.forms import DocumentForm

# Models
from advisers.models import Advisers
from agreements.models import Agreements,AgreementInvestors, AgreementDocuments
from catalogues.models import Banks, CreditStatus, Identifications, Relationships
from clients.models import Clients
from credits.fimubac import views as fimubac
from credits.models import (CreditApplications,
							CreditApplicationsReferences,
							CatEmployeeTypes,
							CreditApplicationJobs,
							CatDocuments,
							CreditsDocuments,
							CreditApplicationPartnership)
from offices.models import BranchOffices
from organizations.models import Organization, OrganizationAddress, OrganizationContacts
from products.models import Products, ProductsRequirements, ProductsInvestor
from amortization.models import AmortizationTable


@user_passes_test(lambda u: u.has_perm('credits.view_creditapplications'), login_url='/error', redirect_field_name=None)
def index(request):
	credits = CreditApplications.objects.all().order_by('id')
	query = request.GET.get('q')
	page = request.GET.get('page')
	results = request.GET.get('results', 10) or 10

	if query:
		credits_tot = CreditApplications.objects.annotate(nombre_completo=Concat('client_id__name', Value(' '), 
																				'client_id__father_lastname', Value(' '), 
																				'client_id__mother_lastname'))
		credits = credits_tot.filter(
				Q(id__icontains=query) | Q(client_id__name__icontains=query) | 
				Q(client_id__father_lastname__icontains=query) | Q(client_id__mother_lastname__icontains=query) | 
				Q(nombre_completo__icontains=query) | Q(folio__icontains=query) | 
				Q(bank__short_name__icontains=query) | Q(investors_app_id__icontains=query)
			).order_by('id')
	
	paginator = Paginator(credits, per_page=results, allow_empty_first_page=True, orphans=5)
	
	try:
		credits = paginator.page(page)
	except PageNotAnInteger:
		credits = paginator.page(1)
	except EmptyPage:
		credits = paginator.page(paginator.num_pages)

	return render(request, 'credits/index.html', {'credits': credits})


# @user_passes_test(lambda u: u.has_perm('credits.add_creditapplications'), login_url='/error', redirect_field_name=None)
# def create(request):
# 	return render(request, 'credits/actions/create.html')


@user_passes_test(lambda u: u.has_perm('credits.add_creditapplications'), login_url='/error', redirect_field_name=None)
def store(request):
	if request.method == 'POST':
		data = request.POST.copy()

		if data.get('client'):
			client_id_data		= data.get('client')
			""" CreditApplications """
			#person_type = data.get('person_type')
			product_id_fimubac	= data.get('products_credits')

			""" CatEmployeeTypes """
			name		= data.get('name')
			description	= data.get('description') 

			credit_applications = CreditApplications.objects.create(
				client_id				= Clients.objects.get(id=data.get('client')),
				loan_amount				= data.get('loan_amount'),
				total_loan				= data.get('total_loan'),
				payment					= data.get('payment'),
				payment_periodicity		= 'Quincenal',
				auth_date				= data.get('auth_date'),
				bank					= Banks.objects.get(bank=data.get('bank')),
				identification			= Identifications.objects.get(identification=data.get('identification')),
				id_number 				= data.get('id_number'),
				clabe					= data.get('clabe'),
				auth_by					= request.user,
				created_by				= request.user,
				status_id				= CreditStatus.objects.get(status=1),
				agreement				= Agreements.objects.get(id=data.get('agreements_credits')),
				product_id				= Products.objects.get(id=data.get('products_credits')),
				branch_office 			= BranchOffices.objects.get(id=data.get('branch_office')),
				adviser 				= Advisers.objects.get(id=data.get('adviser')),
				folio					= data.get('num_application'),
				collection 				= data.get('num_collection')
			)

			credit_references1 = CreditApplicationsReferences.objects.create(
				credit_id				= credit_applications,
				first_name				= data.get('ref1_name'),
				father_lastname			= data.get('ref1_father_lastname'),
				mother_lastname			= data.get('ref1_mother_lastname'),
				phone_contact			= data.get('ref1_phone_contact'),
				relationship			= Relationships.objects.get(relationship=data.get('ref1_relationship')),
				years_of_relationship	= data.get('ref1_years'),
				months_of_relationship	= data.get('ref1_months'),
				created_by				= request.user,
			)

			credit_references2 = CreditApplicationsReferences.objects.create(
				credit_id				= credit_applications,
				first_name				= data.get('ref2_name'),
				father_lastname			= data.get('ref2_father_lastname'),
				mother_lastname			= data.get('ref2_mother_lastname'),
				phone_contact			= data.get('ref2_phone_contact'),
				relationship			= Relationships.objects.get(relationship=data.get('ref2_relationship')),
				years_of_relationship	= data.get('ref2_years'),
				months_of_relationship	= data.get('ref2_months'),
				created_by				= request.user,
			)

			employee_jobs = CreditApplicationJobs.objects.create(
				credit_id			= credit_applications,
				position			= data.get('position'),
				admission_date		= data.get('admission_date'),
				#years_at_work		=years_at_work,
				#months_at_work		=months_at_work,
				month_salary		= data.get('month_salary'),
				workplace			= data.get('workplace'),
				phone_contact		= data.get('phone_contact'),
				employee_number		= data.get('emp_number'),
				created_by			= request.user,
				employee_type_id	= CatEmployeeTypes.objects.get(id=data.get('employee_type')),
				rfc_workplace 		= data.get('rfc_workplace'),
			)

			try:
				partnership = CreditApplicationPartnership.objects.create(
					credit_id		= credit_applications,
					first_name 		= data.get('partner_name'),
					father_lastname	= data.get('partner_father_lastname'),
					mother_lastname	= data.get('partner_mother_lastname'),
					birthdate		= data.get('partner_birthdate'),
					nationality		= data.get('partner_nationality'),
					phone_contact	= None,
					created_by		= request.user,
				)
			except:
				pass

			#Call fimubac ws
			return fimubac.send_application(request, client_id_data, credit_applications, product_id_fimubac, credit_references1, credit_references2)
		else:
			return JsonResponse({'message':'No se cuenta con la información del cliente'}, status=422)

		return render(request, 'credits/actions/create.html')

@user_passes_test(lambda u: u.has_perm('credits.change_creditapplications'), login_url='/error', redirect_field_name=None)
def edit(request, credit_app_id):
	#Get credit app object
	credit = CreditApplications.objects.get(id=credit_app_id)

	#get client obj
	client = Clients.objects.get(id=credit.client_id.id)

	employee_types 	= CatEmployeeTypes.objects.all()
	banks 			= Banks.objects.all().order_by('short_name')
	relationships 	= Relationships.objects.all()

	#Create obj
	obj = {'client': client, 'credit': credit, 'employee_types':employee_types, 'banks':banks, 'relationships':relationships}

	try:
		#Get client job from client applications jobs
		job = CreditApplicationJobs.objects.get(credit_id=credit)
		obj.update({'job':job})
	except:
		job = None
	
	try:
		#get partner from client
		partner = CreditApplicationPartnership.objects.get(credit_id=credit)
		obj.update({'partner':partner})
	except:
		partner = None

	#Get references
	references = CreditApplicationsReferences.objects.filter(credit_id=credit).order_by('id')
	try:
		reference_1 = references[0]
		print('1',reference_1)
		obj.update({'reference_1':reference_1})
	except:
		reference_1 = None

	try:
		reference_2 = references[1]
		print('2',reference_2)
		obj.update({'reference_2':reference_2})
	except:
		reference_2 = None

	#transform Dates into valid fields
	if credit.auth_date:
		auth_credit_date = credit.auth_date.strftime("%Y-%m-%d")
		obj['credit_date'] = auth_credit_date

	if job is not None and job.admission_date:
		job_admission_date = job.admission_date.strftime("%Y-%m-%d")
		obj['job_date'] = job_admission_date
	
	if partner is not None and partner.birthdate:
		partner_birthdate = partner.birthdate.strftime("%Y-%m-%d")
		obj['partner_birthdate'] = partner_birthdate

	return render(request, 'credits/actions/edit.html', obj)

@user_passes_test(lambda u: u.has_perm('credits.change_creditapplications'), login_url='/error', redirect_field_name=None)
def update(request, credit_app_id):
	if request.method == 'POST':
		data = request.POST.copy()

		#Get credit app object
		credit = CreditApplications.objects.get(id=credit_app_id)

		#Get client job from client applications jobs
		job = CreditApplicationJobs.objects.get(credit_id=credit)
		job.workplace				= data.get('workplace')
		job.position				= data.get('position')
		job.phone_contact			= data.get('phone_contact')
		job.admission_date			= data.get('admission_date')
		job.employee_number			= data.get('employee_number')
		job.month_salary			= data.get('month_salary')
		job.save()

		references = CreditApplicationsReferences.objects.filter(credit_id=credit).order_by('id')

		#Get reference 1
		reference_1 = references[0]
		reference_1.first_name				= data.get('ref1_name')
		reference_1.father_lastname			= data.get('ref1_father_lastname')
		reference_1.mother_lastname			= data.get('ref1_mother_lastname')
		reference_1.relationship 			= Relationships.objects.get(relationship=data.get('ref1_relationship'))
		reference_1.phone_contact			= data.get('ref1_phone_contact')
		reference_1.years_of_relationship	= data.get('ref1_years')
		reference_1.months_of_relationship 	= data.get('ref1_months')
		reference_1.save()

		#Get reference 2
		reference_2 = references[1]
		reference_2.first_name				= data.get('ref2_name')
		reference_2.father_lastname			= data.get('ref2_father_lastname')
		reference_2.mother_lastname			= data.get('ref2_mother_lastname')
		reference_2.relationship 			= Relationships.objects.get(relationship=data.get('ref2_relationship'))
		reference_2.phone_contact			= data.get('ref2_phone_contact')
		reference_2.years_of_relationship	= data.get('ref2_years')
		reference_2.months_of_relationship 	= data.get('ref2_months')
		reference_2.save()

		#get partner from client
		try:
			partner = CreditApplicationPartnership.objects.get(credit_id=credit)
			partner.first_name			= data.get('partner_name')
			partner.father_lastname		= data.get('partner_father_lastname')
			partner.mother_lastname		= data.get('partner_mother_lastname')
			partner.birthdate			= data.get('partner_birthdate')
			partner.nationality			= data.get('partner_nationality')
			partner.save()
		except CreditApplicationPartnership.DoesNotExist:
			pass

		return HttpResponse(credit_app_id, status=200)


def goto_credit(request, client_id):

	if isinstance(client_id, dict):
		#print('tas loco crac, es un dict')
		#print(client_id)
		client_obj			= client_id
		get_client_id		= int(str(client_obj['credit_app']).split(' ')[1])
		latest_application	= CreditApplications.objects.get(id=get_client_id)
		client				= Clients.objects.get(id=int(client_obj['client']))
	else:
		#print('normally called')
		client = Clients.objects.get(id=client_id)

	organizations	= Organization.objects.all()
	documents		= CatDocuments.objects.all()
	#onlyfiles = [f for f in listdir('files/') if isfile(join('files/', f))]
	#fs = Storage().listdir('files/')
	#print(onlyfiles)
	c = CatEmployeeTypes.objects.all()

	obj = {'client': client, 'organizations_credits': organizations, 'emp': c, 'documents': documents, 'latest_c': latest_application}

	if isinstance(client_id, dict):
		obj['credit_app']	= client_obj['credit_app']
		obj['saved']		= client_obj['saved']
		obj['result']		= client_obj['result']

	print(obj)
	#return render(request, 'credits/actions/documents.html', obj)
	#html = render_to_string('credits/actions/documents.html', obj, request)
	#return HttpResponse(html)
	return HttpResponse(client_obj['credit_app'])


@user_passes_test(lambda u: u.has_perm('credits.add_creditapplications'), login_url='/error', redirect_field_name=None)
def create_folder(request, credit_app):
	print(credit_app)
	latest_application	= CreditApplications.objects.get(id=credit_app)
	client 				= Clients.objects.get(id=latest_application.client_id.id)
	documents			= CatDocuments.objects.all()
	result				= {'AppId': latest_application.investors_app_id, 'ErrorCode':0}
	obj = {'documents': documents, 'latest_c': latest_application, 'result':result, 'client': client}
	return render(request, 'credits/actions/documents.html', obj)


@user_passes_test(lambda u: u.has_perm('credits.add_creditsdocuments'), login_url='/error', redirect_field_name=None)
def create_file(request):
	""" Method to store the document's of credit application """
	""" if request.method == 'POST':
		form = DocumentForm(request.POST, request.FILES)
		if form.is_valid():
			import pdb; pdb.set_trace()
			form.save()
			return redirect('credits:file')
	else:
		form = DocumentForm()
	return render(request, 'credits/actions/create_file.html', {
		'form': form
	})"""
	#credit_applications = CreditApplications.objects.all()
	documents = CatDocuments.objects.all()
	if (request.method == 'POST'):
		myfile = request.FILES['file']
		#Get document ID from hidden input form
		document_id = request.POST.get('document-obj')
		#Get last credit application id
		latest_c 	= request.POST.get('latest-record')
		#Get client id
		client_id 	= request.POST.get('client-id')
		#Get file type
		file_type 	= myfile.name.split('.')[(len(myfile.name.split('.')) - 1)]

		filepath 	= ('{}{}{}{}{}{}{}'.format((int(latest_c)), '/', (int(latest_c)), '_', document_id, '.', file_type))

		fs 					= FileSystemStorage()
		filename 			= fs.save(filepath, myfile)
		uploaded_file_url 	= fs.url(filename)
		#onlyfiles = [f for f in listdir('{}{}'.format('files/', latest_c)) if isfile(join('{}{}'.format('files/', latest_c), f))]

		credit_app = CreditsDocuments.objects.create(
			credit_id 		= CreditApplications.objects.get(id=latest_c),
			document_id 	= CatDocuments.objects.get(id=document_id),
			path 			= filepath,
			created_by 		= request.user
		)
		print(credit_app)

		return HttpResponse('success')


@user_passes_test(lambda u: u.has_perm('credits.view_creditapplications'), login_url='/error', redirect_field_name=None)
def detail(request):
	credits = CreditApplications.objects.all()
	obj = {'credits': credits}
	return render(request, 'credits/actions/view.html', obj)


@user_passes_test(lambda u: u.has_perm('credits.add_creditsdocuments'), login_url='/error', redirect_field_name=None)
def goto_create_file(request, exp):
	credit_app_id = CreditApplications.objects.get(id=exp)
	documents = CatDocuments.objects.all()

	obj = {'credit_app': credit_app_id, 'documents': documents}
	return render(request, 'credits/actions/create_file.html', obj)


@user_passes_test(lambda u: u.has_perm('credits.view_creditapplications'), login_url='/error', redirect_field_name=None)
def view_records(request):
	# Get all clients Method
	clients = Clients.objects.all().order_by('id')
	credit_document = CreditsDocuments.objects.filter(credit_id__client_id__in=clients).values_list('credit_id__client_id', flat=True).distinct()
	
	#Convert Queryset to List
	credit_list = list(credit_document)

	query = request.GET.get('q')
	page = request.GET.get('page')
	results = request.GET.get('results', 10) or 10

	if query:
		clients_tot = Clients.objects.annotate(nombre_completo=Concat('name', Value(' '), 
																		'father_lastname', Value(' '), 
																		'mother_lastname'))

		clients = clients_tot.filter(
				Q(id__icontains=query) | Q(name__icontains=query) |
				Q(father_lastname__icontains=query) | Q(mother_lastname__icontains=query) | 
				Q(nombre_completo__icontains=query) | Q(RFC__icontains=query) | Q(CURP__icontains=query)
			).distinct().order_by('id')

	paginator = Paginator(clients, per_page=results, allow_empty_first_page=True, orphans=5)

	try:
		clients = paginator.page(page)
	except PageNotAnInteger:
		clients = paginator.page(1)
	except EmptyPage:
		clients = paginator.page(paginator.num_pages)

	#create context for template
	obj = {'clients': clients, 'docs': credit_list}

	return render(request, 'credits/actions/records.html', obj)


@user_passes_test(lambda u: u.has_perm('credits.view_creditapplications'), login_url='/error', redirect_field_name=None)
def record_detail(request, client_id):
	client 		= Clients.objects.get(id=client_id)
	credit_app 	= CreditApplications.objects.filter(client_id=client).order_by('id')

	obj = {'client': client, 'credit_apps': credit_app }
	return render(request, 'credits/actions/record_detail.html', obj)


@user_passes_test(lambda u: u.has_perm('credits.view_creditapplications'), login_url='/error', redirect_field_name=None)
def check_documents(request):
	# Get creditApp ID
	credit_app_id 	= request.POST.get('client_id')
	credit_app 		= CreditApplications.objects.get(id=credit_app_id)
	documents 		= CreditsDocuments.objects.filter(credit_id=credit_app)

	# convert queryset to iterable
	json_documents = list(documents.values())

	return HttpResponse(json.dumps(json_documents, indent=4, sort_keys=True, default=str))


@user_passes_test(lambda u: u.has_perm('credits.view_creditapplications'), login_url='/error', redirect_field_name=None)
def open_file(request):
	if(request.method == 'POST'):
		# Get credit application ID form post request
		credit_app_id 	= request.POST.get('credit_app_id')
		# fromm settings.py get media root path
		fileroot 		= settings.MEDIA_ROOT
		# Create full filepath
		filepath 		= '{}{}{}'.format(fileroot, '/', credit_app_id)
		# Open explorer window with files
		os.startfile(filepath)

	return HttpResponse('success')


@user_passes_test(lambda u: u.has_perm('credits.view_creditapplications'), login_url='/error', redirect_field_name=None)
def export_pdf(request):
	if(request.method == 'POST'):
		# Get list of documents id
		documents_pdf = request.POST.get('documents_pdf')

		#get credit application ID
		credit_app_id = request.POST.get('credit_app_id')

		# fromm settings.py get media root path
		fileroot = settings.MEDIA_ROOT

		# Create full filepath
		filepath = '{}{}{}'.format(fileroot, '/', credit_app_id)

		#Split list of document ID's
		split_document_id = documents_pdf.split(',')

		#path to save PDF
		pdfpath = '{}{}{}{}{}{}'.format(fileroot, '/', credit_app_id, '/', credit_app_id, '.pdf')

		# Array to save full length path
		pdf_documents = []

		# Array to save full length path
		pdf_documents = []

		#Opened images
		image_bits = []

		#file paths
		image_paths = []

		#create dict with splitted values
		for value in range(len(split_document_id)):
			full_doc_name = '{}{}{}{}{}{}'.format(credit_app_id, '/', credit_app_id, '_', split_document_id[value], '.jpg')
			pdf_documents.append(full_doc_name)

		#open list of images
		for document in range(len(pdf_documents)):
			opened_img = Image.open('{}{}{}'.format(fileroot, '/', pdf_documents[document]))
			image_bits.append(opened_img)

			#close image
			opened_img.close()

		for value in range(len(image_bits)):
			image_paths.append(image_bits[value].filename)
		#specify paper size
		letter_size 	= (img2pdf.mm_to_pt(215.9),img2pdf.mm_to_pt(279.4))
		#set paper size varibale to img2pdf
		layout_fun 		= img2pdf.get_layout_fun(letter_size)
		#convert images
		final_pdf 		= img2pdf.convert(image_paths, layout_fun=layout_fun)
		#Open or create pdf file
		file 			= open(pdfpath, 'wb')
		#write pdf files
		file.write(final_pdf)
		#close file
		file.close()

	return HttpResponse('success')


@user_passes_test(lambda u: u.has_perm('credits.view_creditapplications'), login_url='/error', redirect_field_name=None)
def download_file(request, cred_id):
	# from settings.py get media root path
	fileroot 	= settings.MEDIA_ROOT
	# Create full filepath
	filepath 	= '{}{}{}'.format(fileroot, '/', cred_id)
	#path to save PDF
	pdfpath 	= '{}{}{}{}{}{}'.format(fileroot, '/', cred_id, '/', cred_id, '.pdf')
	final_pdf 	= open(pdfpath, 'rb')
	response 	= HttpResponse(final_pdf.read(), content_type='application/pdf')
	response['Content-Disposition'] = '{}{}{}'.format('attachment; filename=', cred_id, '_Expediente.pdf')
	
	return response


def my_credit_detail(request, credit_id):

	#Get Credit application object
	credit_obj = CreditApplications.objects.get(id=credit_id)

	obj = {'credit': credit_obj}

	return render(request, 'credits/actions/credit_detail_client.html', obj)


def measurer(request):
	#Get all clients
	clients 		= Clients.objects.filter(is_active=True)
	#Get all organizations
	organizations 	= Organization.objects.filter(is_active=True)
	#Get all agreements
	agreements 		= Agreements.objects.filter(is_active=True)

	obj = {"clients": len(clients), "organizations_len": len(organizations), "organizations": organizations, "agreements": len(agreements)}

	return render(request, 'credits/actions/measurer.html', obj)



def get_min_year_from_agreement(request):
	agreement_id = request.POST.get('id')
	agreement = Agreements.objects.get(id=agreement_id)

	creditos = CreditApplications.objects.filter(agreement=agreement)

	# convert queryset to iterable
	amortization = AmortizationTable.objects.filter(credit_application__in=creditos)\
									.aggregate(Min('fortnightly_number'), Max('fortnightly_number'))
	min_periodo = amortization['fortnightly_number__min']
	max_periodo = amortization['fortnightly_number__max']

	return JsonResponse({'min': min_periodo[0:4], 'max': max_periodo[0:4]})


def get_active_credits(request):
	if request.method == 'POST':
		data = request.POST.copy()

		organization_id = request.POST.get('organization_id')
		agreement_id = request.POST.get('agreement_id')
		
		'''
		Check date_range value:
		true: 'Quincena Especifica'
		false: 'Rango de Quincenas'
		'''
		date_range = request.POST.get('range')

		'''
		"Estados de solicitudes"
		1	Capturada
		2	Validata
		3	Devolucion
		4	Rechazada
		5	Autorizada
		6	Aperturada
		7	Cancelada
		8	Mesa de Control
		10	Por Formalizar
		11	Pre Captura
		12	Analisis
		20	Integrando
		21	Pre Solicitud Web
		22	Revision Especial
		'''

		#[first_fortnight_day, last_fortnight_day, month]
		fiscal_fortnights = {
			'1': ['01', '14', '01'],
			'2': ['15', '31', '01'],
			'3': ['01', '14', '02'],
			'4': ['15', '28', '02'],
			'5': ['01', '14', '03'],
			'6': ['15', '31', '03'],
			'7': ['01', '14', '04'],
			'8': ['15', '30', '04'],
			'9': ['01', '14', '05'],
			'10': ['15', '31', '05'],
			'11': ['01', '14', '06'],
			'12': ['15', '30', '06'],
			'13': ['01', '14', '07'],
			'14': ['15', '31', '07'],
			'15': ['01', '14', '08'],
			'16': ['15', '31', '08'],
			'17': ['01', '14', '09'],
			'18': ['15', '30', '09'],
			'19': ['01', '14', '10'],
			'20': ['15', '31', '10'],
			'21': ['01', '14', '11'],
			'22': ['15', '30', '11'],
			'23': ['01', '14', '12'],
			'24': ['15', '31', '12'],
		}

		#call for workbook instance
		wb = Workbook()
		ws = wb.active
		
		cols1 = [
			'N° Crédito',
			'RFC',
			'Cliente',
			'Fondeador',
			'Convenio',
			'N° Quincena',
			'Pago Esperado',
			'Pago Realizado',
			'Tipo de Pago'
		]

		cols2 = [
			'N° Crédito',
			'RFC',
			'Pago Esperado',
			'Cliente',
			'Pago Realizado',
			'Diferencia',
			'Comentario',
			'Quincena',
			'Convenio',
			'N° de Pagos'
		]

		#excel rows name
		row_name= [
			'A',
			'B',
			'C',
			'D',
			'E',
			'F',
			'G',
			'H',
			'I',
			'J',
			'K',
			'L',
			'M',
			'N',
			'O',
			'P',
			'Q',
			'R',
			'S',
			'T',
			'U',
			'V',
			'W',
			'X',
			'Y',
			'Z'
		]

		#create col name for xls file
		for col_name in range(len(cols1)):
			ws['{}{}'.format(row_name[col_name], '1')] = cols1[col_name]

	#methods for specific date range
	if (date_range == 'false'):
		specific_date = request.POST.get('specific_date')
		'''
		Split specific_date to obtain desired value
		example format '2018 1'
		the fisrt value is the year and the second, is the fortnight number
		'''
		specific_year = specific_date.split(' ')[0]
		specific_fortnight = specific_date.split(' ')[1]
		fortnight = "%s%s"%(specific_year,str(specific_fortnight).rjust(2,'0'))
		
		#create str that can be converted into date
		date_to_convert = '{}{}{}{}{}'.format(specific_year, '-', fiscal_fortnights[str(int(specific_fortnight))][2], '-', fiscal_fortnights[str(int(specific_fortnight))][0])
		date_to_convert_range = '{}{}{}{}{}'.format(specific_year, '-', fiscal_fortnights[str(int(specific_fortnight))][2], '-', fiscal_fortnights[str(int(specific_fortnight))][1])

		#convert naive date to aware date
		converted_specific_date = datetime.strptime(date_to_convert, "%Y-%m-%d").replace(tzinfo=timezone.utc)
		converted_specific_date_range = datetime.strptime(date_to_convert_range, "%Y-%m-%d").replace(tzinfo=timezone.utc)
		
		#make query sets for specific date
		#get agreement obj
		agreement_obj = Agreements.objects.get(id=agreement_id)
		applications = CreditApplications.objects.filter(agreement=agreement_obj).exclude(status_id__status=7)
		#get active credit applications
		
		# all_applications = CreditApplications.objects.filter(agreement=agreement_obj).filter(created_at__range=[converted_specific_date, converted_specific_date_range])
		all_applications = AmortizationTable.objects.filter(credit_application__in=applications, fortnightly_number=fortnight)
		pending_applications = CreditApplications.objects.filter(agreement=agreement_obj, status_id='1').filter(created_at__range=[converted_specific_date, converted_specific_date_range])
		approved_applications = CreditApplications.objects.filter(agreement=agreement_obj, status_id='5').filter(created_at__range=[converted_specific_date, converted_specific_date_range])
		denied_applications = CreditApplications.objects.filter(agreement=agreement_obj, status_id='4').filter(created_at__range=[converted_specific_date, converted_specific_date_range])

		#get clients data
		clients = Clients.objects.filter(created_at__range=[converted_specific_date, converted_specific_date_range])

		#check amortization table registers for 'fortnightly_payment' in range and sum all registers
		amort_expected_income = AmortizationTable.objects.filter(credit_application__in=applications, fortnightly_number=fortnight).aggregate(Sum('fortnightly_payment'))

		#check amortization table registers for 'amount_paid' in range and sum all registers
		amort_actual_income = AmortizationTable.objects.filter(credit_application__in=applications, fortnightly_number=fortnight).aggregate(Sum('amount_paid'))

		#create obj for xls file
		amort = AmortizationTable.objects.filter(fortnightly_number='{}{}'.format(specific_year, specific_fortnight))

		#create id for filename
		filename_id = '{}{}'.format(specific_year, specific_fortnight)

		#check if querysets have 'None' Values
		if (amort_expected_income['fortnightly_payment__sum'] == None):
			amort_expected_income['fortnightly_payment__sum'] = 0

		if (amort_actual_income['amount_paid__sum'] == None):
			amort_actual_income['amount_paid__sum'] = 0

		#iterate to create xls file contents
		row = 2
		for item in amort:
			if(item.amount_paid == None):
				item.amount_paid = 0

			#get 'cero pagos'
			if(float(item.amount_paid) == 0):
				#credit number
				ws['{}{}'.format('A', str(row))] = str(item.credit_application.id)

				#RFC
				ws['{}{}'.format('B', str(row))] = item.credit_application.client_id.RFC

				#client name
				ws['{}{}'.format('C', str(row))] = '{}{}{}{}{}'.format(item.credit_application.client_id.name, ' ', item.credit_application.client_id.father_lastname, ' ', item.credit_application.client_id.mother_lastname)

				#investor
				ws['{}{}'.format('D', str(row))] = 0 #ProductsInvestor.objects.get(product=item.credit_application.product_id).investor.name

				#agreement
				ws['{}{}'.format('E', str(row))] = item.credit_application.agreement.agreement_name

				#fortnightly_number
				ws['{}{}'.format('F', str(row))] = item.fortnightly_number

				#Pago esperado
				ws['{}{}'.format('G', str(row))] = str(item.fortnightly_payment)

				#Pago Realizado
				ws['{}{}'.format('H', str(row))] = str(item.amount_paid)

				#payment type
				ws['{}{}'.format('I', str(row))] = 'Cero Pago'
			
			#get 'bajo pago'
			elif(float(item.amount_paid) < float(item.fortnightly_payment) and float(item.amount_paid) > 0):
				#credit number
				ws['{}{}'.format('A', str(row))] = str(item.credit_application.id)

				#RFC
				ws['{}{}'.format('B', str(row))] = item.credit_application.client_id.RFC

				#client name
				ws['{}{}'.format('C', str(row))] = '{}{}{}{}{}'.format(item.credit_application.client_id.name, ' ', item.credit_application.client_id.father_lastname, ' ', item.credit_application.client_id.mother_lastname)

				#investor
				ws['{}{}'.format('D', str(row))] = 0 #ProductsInvestor.objects.get(product=item.credit_application.product_id).investor.name
				
				#agreement
				ws['{}{}'.format('E', str(row))] = item.credit_application.agreement.agreement_name

				#fortnightly_number
				ws['{}{}'.format('F', str(row))] = item.fortnightly_number

				#Pago esperado
				ws['{}{}'.format('G', str(row))] = str(item.fortnightly_payment)
				
				#Pago Realizado
				ws['{}{}'.format('H', str(row))] = str(item.amount_paid)

				#payment type
				ws['{}{}'.format('I', str(row))] = 'Bajo Pago'

				print(item.credit_application.investors_app_id)

			#get 'pago normal'
			elif(float(item.fortnightly_payment) == float(item.amount_paid)):
				#credit number
				ws['{}{}'.format('A', str(row))] = str(item.credit_application.id)

				#RFC
				ws['{}{}'.format('B', str(row))] = item.credit_application.client_id.RFC

				#client name
				ws['{}{}'.format('C', str(row))] = '{}{}{}{}{}'.format(item.credit_application.client_id.name, ' ', item.credit_application.client_id.father_lastname, ' ', item.credit_application.client_id.mother_lastname)
				
				#investor
				ws['{}{}'.format('D', str(row))] = 0 #ProductsInvestor.objects.get(product=item.credit_application.product_id).investor.name
				
				#agreement
				ws['{}{}'.format('E', str(row))] = item.credit_application.agreement.agreement_name

				#fortnightly_number
				ws['{}{}'.format('F', str(row))] = item.fortnightly_number

				#Pago esperado
				ws['{}{}'.format('G', str(row))] = str(item.fortnightly_payment)

				#Pago Realizado
				ws['{}{}'.format('H', str(row))] = str(item.amount_paid)

				#payment type
				ws['{}{}'.format('I', str(row))] = 'Normal'

			row = row + 1

	else:
		from_fortnight = request.POST.get('from_date')
		to_fortnight = request.POST.get('to_date')

		'''
		Split form_date and to_fortnight to obtain desired value
		example format '2018 1'
		the fisrt value is the year and the second, is the fortnight number
		'''
		from_year = from_fortnight.split(' ')[0]
		from_fortnight_value = from_fortnight.split(' ')[1]
		
		to_year = to_fortnight.split(' ')[0]
		to_fortnight_value = to_fortnight.split(' ')[1]


		#convert both dates to string
		#date formar: YYYY-M-D
		from_converted = '{}{}{}{}{}'.format(from_year, '-', fiscal_fortnights[str(int(from_fortnight_value))][2], '-', fiscal_fortnights[str(int(from_fortnight_value))][0])
		to_converted = '{}{}{}{}{}'.format(to_year, '-', fiscal_fortnights[str(int(to_fortnight_value))][2], '-', fiscal_fortnights[str(int(to_fortnight_value))][1])

		#convert naive date to aware date
		from_date = datetime.strptime(from_converted, "%Y-%m-%d").replace(tzinfo=timezone.utc)
		to_date = datetime.strptime(to_converted, "%Y-%m-%d").replace(tzinfo=timezone.utc)

		#get agreement obj
		agreement_obj = Agreements.objects.get(id=agreement_id)

		#get active credit applications
		all_applications = CreditApplications.objects.filter(agreement=agreement_obj).filter(created_at__range=[from_date, to_date])
		pending_applications = CreditApplications.objects.filter(agreement=agreement_obj, status_id='1').filter(created_at__range=[from_date, to_date])
		approved_applications = CreditApplications.objects.filter(agreement=agreement_obj, status_id='5').filter(created_at__range=[from_date, to_date])
		denied_applications = CreditApplications.objects.filter(agreement=agreement_obj, status_id='4').filter(created_at__range=[from_date, to_date])

		#get clients data
		clients = Clients.objects.filter(created_at__range=[from_date, to_date])
		
		#check amortization table registers for 'fortnightly_payment' in range and sum all registers
		amort_range_expected = AmortizationTable.objects.all().filter(fortnightly_number__range=['{}{}'.format(from_year, from_fortnight_value), '{}{}'.format(to_year, to_fortnight_value)]).aggregate(Sum('fortnightly_payment'))

		#check amortization table registers for 'amount_paid' in range and sum all registers
		amort_range_actual = AmortizationTable.objects.all().filter(fortnightly_number__range=['{}{}'.format(from_year, from_fortnight_value), '{}{}'.format(to_year, to_fortnight_value)]).aggregate(Sum('amount_paid'))

		#create obj for xls file
		amort = AmortizationTable.objects.filter(fortnightly_number__range=['{}{}'.format(from_year, from_fortnight_value), '{}{}'.format(to_year, to_fortnight_value)]).order_by('fortnightly_number')

		#create id for filename
		filename_id = '{}{}{}{}{}'.format(from_year, from_fortnight_value, '-', to_year, to_fortnight_value)

		#check if querysets have 'None' Values
		if (amort_range_expected['fortnightly_payment__sum'] == None):
			amort_range_expected['fortnightly_payment__sum'] = 0

		if (amort_range_actual['amount_paid__sum'] == None):
			amort_range_actual['amount_paid__sum'] = 0
		
		#iterate to create xls file contents
		row = 2
		for item in amort:
			if(item.amount_paid == None):
				item.amount_paid = 0

			#get 'cero pagos'
			if(float(item.amount_paid) == 0):
				#credit number
				ws['{}{}'.format('A', str(row))] = str(item.credit_application.id)

				#RFC
				ws['{}{}'.format('B', str(row))] = item.credit_application.client_id.RFC

				#client name
				ws['{}{}'.format('C', str(row))] = '{}{}{}{}{}'.format(item.credit_application.client_id.name, ' ', item.credit_application.client_id.father_lastname, ' ', item.credit_application.client_id.mother_lastname)

				#investor
				ws['{}{}'.format('D', str(row))] = ProductsInvestor.objects.get(product=item.credit_application.product_id).investor.name
				
				#agreement
				ws['{}{}'.format('E', str(row))] = item.credit_application.agreement.agreement_name

				#fortnightly_number
				ws['{}{}'.format('F', str(row))] = item.fortnightly_number

				#Pago esperado
				ws['{}{}'.format('G', str(row))] = str(item.fortnightly_payment)

				#Pago Realizado
				ws['{}{}'.format('H', str(row))] = str(item.amount_paid)

				#payment type
				ws['{}{}'.format('I', str(row))] = 'Cero Pago'
			
			#get 'bajo pago'
			elif(float(item.amount_paid) < float(item.fortnightly_payment) and float(item.amount_paid) > 0):
				#credit number
				ws['{}{}'.format('A', str(row))] = str(item.credit_application.id)

				#RFC
				ws['{}{}'.format('B', str(row))] = item.credit_application.client_id.RFC

				#client name
				ws['{}{}'.format('C', str(row))] = '{}{}{}{}{}'.format(item.credit_application.client_id.name, ' ', item.credit_application.client_id.father_lastname, ' ', item.credit_application.client_id.mother_lastname)

				#investor
				ws['{}{}'.format('D', str(row))] = ProductsInvestor.objects.get(product=item.credit_application.product_id).investor.name
				
				#agreement
				ws['{}{}'.format('E', str(row))] = item.credit_application.agreement.agreement_name

				#fortnightly_number
				ws['{}{}'.format('F', str(row))] = item.fortnightly_number

				#Pago esperado
				ws['{}{}'.format('G', str(row))] = str(item.fortnightly_payment)
				
				#Pago Realizado
				ws['{}{}'.format('H', str(row))] = str(item.amount_paid)

				#payment type
				ws['{}{}'.format('I', str(row))] = 'Bajo Pago'

				print(item.credit_application.investors_app_id)

			#get 'pago normal'
			elif(float(item.fortnightly_payment) == float(item.amount_paid)):
				#credit number
				ws['{}{}'.format('A', str(row))] = str(item.credit_application.id)

				#RFC
				ws['{}{}'.format('B', str(row))] = item.credit_application.client_id.RFC

				#client name
				ws['{}{}'.format('C', str(row))] = '{}{}{}{}{}'.format(item.credit_application.client_id.name, ' ', item.credit_application.client_id.father_lastname, ' ', item.credit_application.client_id.mother_lastname)
				
				#investor
				ws['{}{}'.format('D', str(row))] = ProductsInvestor.objects.get(product=item.credit_application.product_id).investor.name
				
				#agreement
				ws['{}{}'.format('E', str(row))] = item.credit_application.agreement.agreement_name

				#fortnightly_number
				ws['{}{}'.format('F', str(row))] = item.fortnightly_number

				#Pago esperado
				ws['{}{}'.format('G', str(row))] = str(item.fortnightly_payment)

				#Pago Realizado
				ws['{}{}'.format('H', str(row))] = str(item.amount_paid)

				#payment type
				ws['{}{}'.format('I', str(row))] = 'Normal'

			row = row + 1

	#create .xlsx file to download
	#create path to save document
	fileroot = settings.MEDIA_ROOT

	#Check if measurer_files exists
	try:
		os.mkdir(os.path.join(fileroot, 'measurer_files'))
	except OSError as error:
		print(error)

	#create full filepath to store layouts
	filepath = '{}{}{}{}'.format(fileroot, '/', 'measurer_files', '/')

	#save document?
	wb.save('{}{}{}{}'.format(filepath, 'Detalle de Pagos', filename_id, '.xlsx'))
		

	#create dict to store all data, then pass it to context or as a response
	measurer_obj = {}

	#push number of new clients into 
	measurer_obj['total_credits'] = len(all_applications)
	measurer_obj['new_clients'] = len(clients)
	measurer_obj['pending_credits'] = len(pending_applications)
	measurer_obj['approved_credits'] = len(approved_applications)
	measurer_obj['denied_credits'] = len(denied_applications)
	measurer_obj['filename'] = '{}{}{}'.format('Detalle de Pagos', filename_id, '.xlsx')
	
	if(date_range == 'false'):
		#convert expected income float to currency-like format
		measurer_obj['expected_income'] = "${:,.2f}".format(float(amort_expected_income['fortnightly_payment__sum']))

		#convert actual income float to currency-like format
		measurer_obj['actual_income'] = "${:,.2f}".format(float(amort_actual_income['amount_paid__sum']))

		#get value of not recovered income
		not_recovered = (float(amort_expected_income['fortnightly_payment__sum']) - float(amort_actual_income['amount_paid__sum']))
		measurer_obj['not_recovered'] = "${:,.2f}".format(not_recovered)
	
	elif (date_range == 'true'):
		#convert expected income float to currency-like format
		measurer_obj['expected_income'] = "${:,.2f}".format(float(amort_range_expected['fortnightly_payment__sum']))

		#convert actual income float to currency-like format
		measurer_obj['actual_income'] = "${:,.2f}".format(float(amort_range_actual['amount_paid__sum']))

		#get value of not recovered income
		not_recovered = (float(amort_range_expected['fortnightly_payment__sum']) - float(amort_range_actual['amount_paid__sum']))
		measurer_obj['not_recovered'] = "${:,.2f}".format(not_recovered)


	#return HttpResponse(json.dumps(list(measurer_obj.values), indent=4, sort_keys=True, default=str))
	return HttpResponse(json.dumps(measurer_obj))

def download_measurer(request, doc):
	
	#init storage variables
	fileroot = settings.MEDIA_ROOT
	
	#file dir
	filepath = '{}{}{}{}{}'.format(fileroot, '/', 'measurer_files', '/', doc)

	#open file to start download
	download_doc = open(filepath, 'rb')

	#create headers to download file
	response = HttpResponse(download_doc.read(), content_type='application/vnd.ms-excel')

	#append attachement and filename to response
	response['Content-Disposition'] = '{}{}'.format('attachement; filename=', doc)

	#close and delete the downloaded file
	download_doc.close()
	os.remove(filepath)

	return response
