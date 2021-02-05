from openpyxl import Workbook
from django.views.generic import *

from django.http.response import HttpResponse


from agreements.models import Agreements
from organizations.models import Organization
from products.models import Products
from clients.models import Clients


class Instalation_Report(TemplateView):

    def get(self, request, *args, **kwargs):

        organization = Organization.objects.all() 
        agreement = Agreements.objects.all()
        product = Products.objects.all()
        client = Clients.objects.all()

        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'Archivo de instalacion'

        ws.merge_cells('B1:E1')
        
        ws['B3'] = 'Version'
        ws['C3'] = 'Dependencia'
        ws['D3'] = 'Nombre'
        ws['E3'] = 'Folio'
        ws['F3'] = 'RFC'
        ws['G3'] = 'Capital'
        ws['H3'] = 'Pagare'
        ws['I3'] = 'Quincenas'
        ws['J3'] = 'Descuento quincenal'
        ws['K3'] = 'Descuento mensual'
        ws['L3'] = 'Clave de descuento'
        ws['M3'] = 'Fondeador'
        ws['N3'] = 'Quincena inicial'
        ws['O3'] = 'Quincena final'
        ws['P3'] = 'Quincena en proceso'
        ws['Q3'] = 'Observaciones'
        ws['R3'] = 'Tipo de movimiento'
        ws['S3'] = 'Pagos'

        count = 4
        for a in client:    
            num = str(count)
            ws['B'+ num] = a.name
            count = count + 1
        
        count = 4
        for a in client:    
            num = str(count)
            ws['C'+ num] = a.RFC
            count = count + 1
        
        count = 4
        for a in client:    
            num = str(count)
            ws['D'+ num] = a.name
            count = count + 1
        
        count = 4
        for a in client:    
            num = str(count)
            ws['E'+ num] = a.RFC
            count = count + 1
        
        count = 4
        for a in client:    
            num = str(count)
            ws['F'+ num] = a.RFC
            count = count + 1



        nombre_archivo = "Instalation.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; finename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response
